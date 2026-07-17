"""Enhanced file tools — PDF reading, image OCR, Excel/CSV processing.

Optional dependencies:
- PDF: pip install pymupdf  (or: pip install cheetahclaws[files])
- OCR: pip install pytesseract Pillow  (+ system tesseract)
- Excel: pip install openpyxl  (or: pip install cheetahclaws[files])
"""
from __future__ import annotations

import csv
import io
from pathlib import Path

from cheetahclaws.tool_registry import ToolDef, register_tool


_DEFAULT_PDF_MAX_FILE_BYTES = 32 * 1024 * 1024
_DEFAULT_SUMMARIZE_MAX_INPUT_BYTES = 16 * 1024 * 1024
_DEFAULT_SUMMARIZE_CHUNK_OUTPUT_CHARS = 8_000
_DEFAULT_SUMMARIZE_REDUCE_INPUT_CHARS = 200_000


def _extract_page_prefix(page, fitz_module, char_cap: int) -> tuple[str, bool]:
    """Extract a page in clipped bands instead of materializing all page text.

    PyMuPDF's plain ``page.get_text()`` builds the complete page string first.
    Reading shallow horizontal bands (and narrower columns for very wide
    pages) keeps peak extraction work bounded even for a pathological one-page
    PDF. Pages without geometry are rejected safely: the compatibility fallback
    of calling ``get_text()`` would materialize the entire page.
    """
    rect = getattr(page, "rect", None)
    if rect is None or not getattr(rect, "height", 0):
        return "", True

    chunks: list[str] = []
    captured = 0
    band_height = min(144.0, max(36.0, float(rect.height) / 16.0))
    max_bands = 256
    page_width = float(getattr(rect, "width", float(rect.x1) - float(rect.x0)))
    tile_width = min(612.0, max(72.0, page_width))
    columns = max(1, min(8, int((page_width + tile_width - 1) // tile_width)))
    max_tiles = 512
    y = float(rect.y0)
    bands_read = 0
    tiles_read = 0
    tile_limit_hit = False
    while (
        y < float(rect.y1)
        and captured < char_cap
        and bands_read < max_bands
        and tiles_read < max_tiles
    ):
        for column in range(columns):
            if captured >= char_cap or tiles_read >= max_tiles:
                tile_limit_hit = tiles_read >= max_tiles
                break
            x0 = float(rect.x0) + column * tile_width
            x1 = min(x0 + tile_width, float(rect.x1))
            clip = fitz_module.Rect(x0, y, x1, min(y + band_height, rect.y1))
            text = page.get_text("text", clip=clip)
            remaining = char_cap - captured
            if len(text) > remaining:
                chunks.append(text[:remaining])
                return "".join(chunks), True
            chunks.append(text)
            captured += len(text)
            tiles_read += 1
        y += band_height
        bands_read += 1
    return "".join(chunks), y < float(rect.y1) or tile_limit_hit


def _read_pdf(params: dict, config: dict) -> str:
    """Read text content from a PDF file."""
    try:
        import fitz  # pymupdf
    except ImportError:
        return (
            "PDF reading requires pymupdf. Install with:\n"
            "  pip install pymupdf\n"
            "Or: pip install cheetahclaws[files]"
        )

    file_path = params["file_path"]
    pages = params.get("pages")  # e.g. "1-5" or "3" or None (all)
    p = Path(file_path)

    if not p.exists():
        return f"Error: file not found: {file_path}"
    if not p.suffix.lower() == ".pdf":
        return f"Error: not a PDF file: {file_path}"

    try:
        try:
            char_cap = int(config.get("pdf_extract_max_chars", 50_000))
        except (TypeError, ValueError):
            char_cap = 50_000
        try:
            page_cap = int(config.get("pdf_extract_max_pages", 50))
        except (TypeError, ValueError):
            page_cap = 50
        try:
            file_byte_cap = int(config.get(
                "pdf_extract_max_file_bytes", _DEFAULT_PDF_MAX_FILE_BYTES,
            ))
        except (TypeError, ValueError):
            file_byte_cap = _DEFAULT_PDF_MAX_FILE_BYTES
        char_cap = max(1_000, char_cap)
        page_cap = max(1, page_cap)
        file_byte_cap = max(1_024, file_byte_cap)
        if p.stat().st_size > file_byte_cap:
            return (
                f"Error: PDF is larger than the {file_byte_cap:,}-byte extraction "
                "limit; use a narrower source file or raise pdf_extract_max_file_bytes."
            )

        doc = fitz.open(str(p))
        try:
            total = len(doc)

            # Parse page range. Explicit page lists are capped too: otherwise
            # a request such as ``1-999999`` allocates and extracts far more
            # than one agent turn can safely use.
            if pages:
                page_list, page_range_truncated = _parse_page_range_capped(
                    pages, total, max_pages=page_cap,
                )
            else:
                page_list = list(range(min(total, page_cap)))
                page_range_truncated = total > page_cap

            text_parts = []
            extracted_chars = 0
            source_truncated = page_range_truncated
            for i in page_list:
                if extracted_chars >= char_cap:
                    source_truncated = True
                    break
                if not (0 <= i < total):
                    continue
                page = doc[i]
                text, page_truncated = _extract_page_prefix(
                    page, fitz, char_cap - extracted_chars,
                )
                source_truncated = source_truncated or page_truncated
                clean_text = text.strip()
                if not clean_text:
                    continue
                remaining = char_cap - extracted_chars
                if len(clean_text) > remaining:
                    clean_text = clean_text[:remaining]
                    source_truncated = True
                text_parts.append(f"--- Page {i+1} ---\n{clean_text}")
                extracted_chars += len(clean_text)
                if extracted_chars >= char_cap:
                    source_truncated = True
                    break
        finally:
            doc.close()

        if not text_parts:
            message = f"PDF has {total} pages but no extractable text (may be scanned/image-only)."
            if source_truncated:
                message += (
                    f" Extraction also stopped at {char_cap:,} characters or "
                    f"{page_cap} pages; use a narrower `pages` range."
                )
            return message

        header = f"PDF: {p.name} ({total} pages, showing {len(text_parts)})\n\n"
        content = "\n\n".join(text_parts)
        full_text = header + content

        # Defense-in-depth: even if the model ignores the
        # research_assistant template's "use SummarizeLargeFile"
        # instruction and calls ReadPDF directly on a huge PDF, the
        # ReadPDF response itself routes the model to the right tool
        # before the raw 70KB+ of PDF text overflows the next API call.
        # See _maybe_redirect_to_summarize for the threshold logic.
        redirect = None if config.get("_skip_summary_redirect") else _maybe_redirect_to_summarize(
            full_text, str(p), config,
        )
        if redirect:
            return redirect

        if source_truncated:
            content += (
                f"\n\n[... ReadPDF stopped at {char_cap:,} extracted characters "
                f"or {page_cap} pages; use a narrower `pages` range or "
                "SummarizeLargeFile for complete coverage ...]"
            )

        return header + content

    except Exception as e:
        return f"Error reading PDF: {type(e).__name__}: {e}"


def _read_image(params: dict, config: dict) -> str:
    """Extract text from an image using OCR."""
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        return (
            "OCR requires pytesseract and Pillow. Install with:\n"
            "  pip install pytesseract Pillow\n"
            "Also install Tesseract OCR engine:\n"
            "  macOS: brew install tesseract\n"
            "  Ubuntu: sudo apt install tesseract-ocr\n"
            "  Windows: https://github.com/UB-Mannheim/tesseract/wiki"
        )

    file_path = params["file_path"]
    lang = params.get("language", "eng")
    p = Path(file_path)

    if not p.exists():
        return f"Error: file not found: {file_path}"

    try:
        img = Image.open(str(p))
        text = pytesseract.image_to_string(img, lang=lang)

        if not text.strip():
            return f"No text detected in image: {p.name}"

        return f"OCR result from {p.name} ({img.size[0]}x{img.size[1]}, lang={lang}):\n\n{text.strip()}"

    except Exception as e:
        return f"OCR error: {type(e).__name__}: {e}"


def ocr_image_bytes(image_bytes: bytes, lang: str = "eng") -> str:
    """Best-effort local OCR on raw image bytes. Returns extracted text or ''.

    Never raises: missing pytesseract/Pillow/tesseract-binary, corrupt bytes,
    or an empty result all collapse to ''. Used by the ``/image`` command to
    enrich the prompt with the verbatim text content of a clipboard image so
    that NON-vision models can still act on screenshots (error dumps, code,
    receipts, tables). Vision models simply get both signals.
    """
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        return ""
    try:
        img = Image.open(io.BytesIO(image_bytes))
        return pytesseract.image_to_string(img, lang=lang).strip()
    except Exception:
        return ""


def _read_excel(params: dict, config: dict) -> str:
    """Read data from Excel (.xlsx) or CSV files."""
    file_path = params["file_path"]
    sheet = params.get("sheet")
    max_rows = min(params.get("max_rows", 100), 500)
    p = Path(file_path)

    if not p.exists():
        return f"Error: file not found: {file_path}"

    ext = p.suffix.lower()

    try:
        if ext == ".csv":
            return _read_csv_file(p, max_rows)
        elif ext in (".xlsx", ".xls"):
            return _read_xlsx_file(p, sheet, max_rows)
        elif ext == ".tsv":
            return _read_csv_file(p, max_rows, delimiter="\t")
        else:
            return f"Unsupported format: {ext}. Supported: .csv, .tsv, .xlsx, .xls"

    except Exception as e:
        return f"Error reading {ext} file: {type(e).__name__}: {e}"


def _read_csv_file(path: Path, max_rows: int, delimiter: str = ",") -> str:
    """Read a CSV/TSV file."""
    with open(path, encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        rows = []
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append(row)

    if not rows:
        return f"Empty file: {path.name}"

    # Format as table
    return _format_table(rows, path.name, total_hint=f"showing {len(rows)} rows")


def _read_xlsx_file(path: Path, sheet: str | None, max_rows: int) -> str:
    """Read an Excel file."""
    try:
        import openpyxl
    except ImportError:
        return (
            "Excel reading requires openpyxl. Install with:\n"
            "  pip install openpyxl\n"
            "Or: pip install cheetahclaws[files]"
        )

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    if sheet and sheet in sheet_names:
        ws = wb[sheet]
    elif sheet and sheet.isdigit():
        idx = int(sheet) - 1
        if 0 <= idx < len(sheet_names):
            ws = wb[sheet_names[idx]]
        else:
            wb.close()
            return f"Sheet index {sheet} out of range. Available: {', '.join(sheet_names)}"
    else:
        ws = wb.active

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        rows.append([str(c) if c is not None else "" for c in row])

    wb.close()

    if not rows:
        return f"Empty sheet: {ws.title}"

    header = f"Excel: {path.name}, sheet: {ws.title}"
    if len(sheet_names) > 1:
        header += f" (sheets: {', '.join(sheet_names)})"

    return _format_table(rows, header, total_hint=f"showing {len(rows)} rows")


def _format_table(rows: list[list], title: str, total_hint: str = "") -> str:
    """Format rows as a readable text table."""
    if not rows:
        return "(empty)"

    # Calculate column widths (cap at 30 chars per col)
    n_cols = max(len(r) for r in rows)
    widths = [0] * n_cols
    for row in rows[:20]:  # sample first 20 rows for width
        for j, cell in enumerate(row):
            if j < n_cols:
                widths[j] = min(max(widths[j], len(str(cell))), 30)

    lines = [f"{title} ({total_hint})\n"]

    for i, row in enumerate(rows):
        cells = []
        for j in range(n_cols):
            val = row[j] if j < len(row) else ""
            cells.append(str(val)[:30].ljust(widths[j]))
        lines.append(" | ".join(cells))
        if i == 0:
            lines.append("-+-".join("-" * w for w in widths))

    return "\n".join(lines)


def _parse_page_range_capped(
    spec: str,
    total: int,
    max_pages: int | None = None,
) -> tuple[list[int], bool]:
    """Parse a page range and state whether a requested page was omitted."""
    pages = []
    seen = set()

    def _add(page: int) -> bool:
        # Ignore out-of-range singleton values just as ranges are clamped.
        # They must not consume the page budget ahead of valid requests.
        if not 0 <= page < total:
            return False
        if page in seen:
            return False
        if max_pages is not None and len(pages) >= max_pages:
            return True
        seen.add(page)
        pages.append(page)
        return False

    for part in spec.split(","):
        part = part.strip()
        if "-" in part:
            a, b = part.split("-", 1)
            start = max(int(a) - 1, 0)
            end = min(int(b), total)
            for page in range(start, end):
                if _add(page):
                    return sorted(pages), True
        elif part.isdigit():
            if _add(int(part) - 1):
                return sorted(pages), True
    return sorted(pages), False


def _parse_page_range(spec: str, total: int, max_pages: int | None = None) -> list[int]:
    """Parse page range like '1-5', '3', '1,3,5-8'."""
    return _parse_page_range_capped(spec, total, max_pages)[0]


# ── Register ─────────────────────────────────────────────────────────────

register_tool(ToolDef(
    name="ReadPDF",
    schema={
        "name": "ReadPDF",
        "description": (
            "Read text content from a PDF file. Extracts text from specified pages. "
            "For scanned/image-only PDFs, use ReadImage with OCR instead."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Absolute path to the PDF file",
                },
                "pages": {
                    "type": "string",
                    "description": "Page range: '1-5', '3', '1,3,5-8'. Default: first 50 pages.",
                },
            },
            "required": ["file_path"],
        },
    },
    func=_read_pdf,
    read_only=True,
    concurrent_safe=True,
))

register_tool(ToolDef(
    name="ReadImage",
    schema={
        "name": "ReadImage",
        "description": (
            "Extract text from an image using OCR (Tesseract). "
            "Supports PNG, JPG, TIFF, BMP. Useful for scanned documents, screenshots, "
            "and image-only PDFs (convert to image first)."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Absolute path to the image file",
                },
                "language": {
                    "type": "string",
                    "description": "OCR language code (default: eng). Examples: chi_sim (Chinese), jpn (Japanese), deu (German)",
                    "default": "eng",
                },
            },
            "required": ["file_path"],
        },
    },
    func=_read_image,
    read_only=True,
    concurrent_safe=True,
))

register_tool(ToolDef(
    name="ReadSpreadsheet",
    schema={
        "name": "ReadSpreadsheet",
        "description": (
            "Read data from Excel (.xlsx/.xls) or CSV/TSV files. "
            "Returns formatted table with column alignment. "
            "For Excel files with multiple sheets, specify the sheet name or number."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Absolute path to the spreadsheet file (.xlsx, .xls, .csv, .tsv)",
                },
                "sheet": {
                    "type": "string",
                    "description": "Sheet name or number (Excel only, default: active sheet)",
                },
                "max_rows": {
                    "type": "integer",
                    "description": "Max rows to return (default: 100, max: 500)",
                    "default": 100,
                },
            },
            "required": ["file_path"],
        },
    },
    func=_read_excel,
    read_only=True,
    concurrent_safe=True,
))


# ── SummarizeLargeFile — multi-agent map-reduce for files that overflow context

_TOKENS_PER_CHAR = 1 / 2.8   # matches compaction.estimate_tokens
_SUMMARIZE_RESERVED_TOKENS = 8500   # system + template + output cap + safety
_SUMMARIZE_MIN_CHUNK_TOKENS = 2000


def _estimate_text_tokens(text: str) -> int:
    """Rough conservative token estimator for plain text. Matches the
    chars/2.8 ratio compaction.estimate_tokens uses."""
    return len(text) if _is_cjk_heavy(text) else int(len(text) * _TOKENS_PER_CHAR)


def _is_cjk_heavy(text: str, sample_chars: int = 2000) -> bool:
    """Heuristic: does this text contain enough CJK that we should use
    1-token-per-char estimation instead of chars/2.8?

    CJK content tokenizes at roughly 1 token per character on most
    tokenizers — vs ~2.8 chars/token for English. A 32K-char "English-
    sized" tool result that's actually Chinese hits 32K tokens, not 11K.
    """
    if not text:
        return False
    sample = text[:sample_chars]
    cjk = sum(
        1 for ch in sample
        if "一" <= ch <= "鿿"   # CJK Unified Ideographs
        or "㐀" <= ch <= "䶿"   # CJK Extension A
        or "぀" <= ch <= "ゟ"   # Hiragana
        or "゠" <= ch <= "ヿ"   # Katakana
        or "가" <= ch <= "힯"   # Hangul
    )
    return cjk / max(len(sample), 1) > 0.20   # ≥20% CJK → treat as CJK-heavy


def _maybe_redirect_to_summarize(text: str, file_path: str,
                                    config: dict) -> str | None:
    """If `text` is too big to safely return as a tool result without
    risking context overflow on the next API call, return a SHORT
    redirect message that tells the model to call SummarizeLargeFile
    instead. Otherwise return None (caller returns the original text).

    This is the deterministic backstop: even if the model ignores the
    template's "use SummarizeLargeFile" instruction and calls Read/
    ReadPDF directly, Read's response itself routes the model to the
    correct tool. The model never sees the overflow-causing raw content.
    """
    if not text:
        return None
    from cheetahclaws.compaction import get_context_limit

    model = config.get("model", "")
    declared_ctx = get_context_limit(model) or 32768
    # ⚠️ Don't blindly trust declared context limits — the `custom/`
    # provider defaults to 128000 in PROVIDERS but the user might be
    # serving a 32K-context model behind it (qwen2.5-72b, llama 3 8B
    # etc.). Cap the ceiling at 30000 so we redirect early enough to
    # protect the smallest commonly-used model. For 200K-context models
    # this is very conservative but harmless: SummarizeLargeFile is
    # always cheap on small files (single-shot path).
    safe_ctx = min(declared_ctx, 30000)

    # Worst-case: assume CJK content tokenizes 1:1 with chars. For pure
    # English content, this is ~3× too conservative — but that's safe
    # (we'd only redirect on truly large files).
    if _is_cjk_heavy(text):
        estimated_tokens = len(text)
    else:
        estimated_tokens = _estimate_text_tokens(text)

    # Reserve ~6K for system prompt + framing + tool schemas + room for
    # the model's response. 70% of remaining is the safe ceiling for
    # any single tool result — beyond that we're risking overflow.
    safe_tool_result_tokens = int((safe_ctx - 6000) * 0.7)
    if estimated_tokens <= safe_tool_result_tokens:
        return None

    active_names = config.get("_active_tool_names")
    if active_names is not None:
        summary_enabled = "SummarizeLargeFile" in active_names
    else:
        profile = config.get("tool_profile")
        # Preserve direct-call compatibility while respecting callers that
        # explicitly request the compact standard profile.
        summary_enabled = profile is None or profile in {"research", "full"}

    # Generate a redirect with a small preview so the model has *some*
    # context to decide on a focus.
    preview_chars = min(1500, len(text) // 8)
    preview = text[:preview_chars].rstrip()
    if not summary_enabled:
        return (
            f"[ReadTooLarge: file `{file_path}` is too large to return "
            f"directly — estimated {estimated_tokens:,} tokens vs model "
            f"context {declared_ctx:,} (safe tool-result ceiling "
            f"{safe_tool_result_tokens:,}).\n\n"
            "USE INSTEAD: Call `Read` again with a narrower `offset` and "
            "`limit`. A document-summary tool is not enabled for this tool "
            "profile.\n\n"
            f"PREVIEW (first {preview_chars} chars, for context only):\n\n"
            f"```\n{preview}\n```]"
        )
    return (
        f"[ReadTooLarge: file `{file_path}` is too large to return "
        f"directly — estimated {estimated_tokens:,} tokens vs model "
        f"context {declared_ctx:,} (safe tool-result ceiling "
        f"{safe_tool_result_tokens:,}).\n\n"
        f"USE INSTEAD: Call **SummarizeLargeFile** with "
        f"`file_path='{file_path}'` (and optional `focus=...`). It will "
        f"chunk the file, summarize each chunk via parallel sub-LLM "
        f"calls, then merge into a single summary that fits in your "
        f"context — no further action needed.\n\n"
        f"PREVIEW (first {preview_chars} chars, for context only — do "
        f"NOT use this as the file's content):\n\n```\n{preview}\n```]"
    )


def _read_file_for_summary(file_path: str, config: dict) -> str:
    """Read content from a file, dispatching to the right reader based on
    extension. Returns the raw text or a string starting with 'Error:' on
    failure."""
    p = Path(file_path)
    if not p.exists():
        return f"Error: file not found: {file_path}"
    if p.is_dir():
        return f"Error: {file_path} is a directory, not a file"
    suffix = p.suffix.lower()
    if suffix == ".pdf":
        # Reuse the bounded PDF extractor but bypass its user-facing redirect:
        # otherwise this summarizer would summarize its own "call
        # SummarizeLargeFile" instruction rather than the PDF text.
        internal = {**config, "_skip_summary_redirect": True}
        content = _read_pdf({"file_path": str(p)}, internal)
        if "[... ReadPDF stopped" in content:
            return (
                "Error: PDF exceeds the configured extraction cap for summarization; "
                "use a narrower `pages` range or raise the PDF extraction limits."
            )
        return content
    # Plain text / code / markdown / etc.
    try:
        try:
            byte_cap = int(config.get(
                "summarize_max_input_bytes", _DEFAULT_SUMMARIZE_MAX_INPUT_BYTES,
            ))
        except (TypeError, ValueError):
            byte_cap = _DEFAULT_SUMMARIZE_MAX_INPUT_BYTES
        byte_cap = max(1_024, byte_cap)
        if p.stat().st_size > byte_cap:
            return (
                f"Error: file exceeds the {byte_cap:,}-byte summary input limit; "
                "use a smaller file or raise summarize_max_input_bytes."
            )
        with p.open("rb") as handle:
            raw = handle.read(byte_cap)
        return raw.decode("utf-8", errors="replace")
    except Exception as e:
        return f"Error reading {file_path}: {type(e).__name__}: {e}"


def _summarize_chunk_via_llm(text: str, focus: str, config: dict,
                               mode: str = "single",
                               chunk_idx: int = 0,
                               total_chunks: int = 0) -> str:
    """Run a single LLM call to summarize one chunk (or merge chunk
    summaries). Uses the session's current model with no_tools=True so
    no recursive tool calls happen.

    Returns the summary text, or an `[error: ...]` marker string on
    failure (so a single chunk failure doesn't sink the whole map-reduce)."""
    from cheetahclaws.providers import stream, TextChunk

    focus_clause = f" Focus on: {focus}." if focus else ""

    if mode == "single":
        sys_msg = (
            "You are summarizing a document for a researcher. Be concrete, "
            "specific, comprehensive. Preserve all named entities, numbers, "
            "and citations."
        )
        user_msg = (
            f"Summarize this document.{focus_clause}\n\n"
            f"Cover: title/author/venue, problem, method, results, "
            f"limitations, connections to related work. Use Markdown headings.\n\n"
            f"{text}"
        )
    elif mode == "map":
        sys_msg = (
            f"You are summarizing chunk {chunk_idx} of {total_chunks} of a "
            f"long document. Each chunk is a contiguous slice; you may not "
            f"see the whole document. Capture EVERY concrete fact, claim, "
            f"number, named entity, and method described in your chunk. A "
            f"later 'reduce' step will merge your summary with the others, "
            f"so prioritize specifics over polish."
        )
        user_msg = (
            f"Summarize chunk {chunk_idx}/{total_chunks}.{focus_clause}\n\n"
            f"Be exhaustive on specifics — names, numbers, results, citations, "
            f"section headings if present. Output as Markdown bullets.\n\n"
            f"{text}"
        )
    else:  # mode == "reduce"
        sys_msg = (
            "You are merging summaries of consecutive chunks of a single "
            "document into ONE unified summary. The chunks were processed "
            "independently; reconcile any tension; deduplicate; preserve "
            "all specifics (numbers, names, citations)."
        )
        user_msg = (
            f"Below are summaries of consecutive sections of a single "
            f"document.{focus_clause}\n\nMerge them into ONE coherent "
            f"summary covering: title/author/venue, problem, method, "
            f"results, limitations, connections to related work. Use "
            f"Markdown headings.\n\n{text}"
        )

    out: list[str] = []
    try:
        output_cap = int(config.get(
            "summarize_chunk_max_output_chars", _DEFAULT_SUMMARIZE_CHUNK_OUTPUT_CHARS,
        ))
    except (TypeError, ValueError):
        output_cap = _DEFAULT_SUMMARIZE_CHUNK_OUTPUT_CHARS
    output_cap = max(500, output_cap)
    output_chars = 0
    internal = {**config, "no_tools": True}
    try:
        for ev in stream(config["model"], sys_msg,
                          [{"role": "user", "content": user_msg}], [],
                          internal):
            if isinstance(ev, TextChunk):
                remaining = output_cap - output_chars
                if remaining <= 0:
                    break
                out.append(ev.text[:remaining])
                output_chars += min(len(ev.text), remaining)
                if len(ev.text) > remaining:
                    break
    except Exception as e:
        return f"[chunk-summarize error: {type(e).__name__}: {str(e)[:200]}]"
    return "".join(out).strip() or "[chunk-summarize: empty response]"


def _plan_chunks(content: str, model_ctx: int) -> list[str]:
    """Split `content` into N chunks each fitting within
    `(model_ctx - reserved) / chars_per_token` chars, with a small overlap
    for continuity. N scales with content size:
        ≤1 chunk-budget       → 1 chunk
        ~3 chunk-budgets      → 3 chunks
        ~10 chunk-budgets     → 10 chunks
        etc. (no hard cap — grows with file)
    """
    cjk_heavy = _is_cjk_heavy(content)
    n_tokens = _estimate_text_tokens(content)
    chunk_token_budget = max(_SUMMARIZE_MIN_CHUNK_TOKENS,
                              model_ctx - _SUMMARIZE_RESERVED_TOKENS)
    if n_tokens <= chunk_token_budget:
        return [content]
    # Compute target char-size per chunk so all chunks roughly equal.
    chars_per_token = 1.0 if cjk_heavy else _TOKENS_PER_CHAR
    chunk_char_budget = int(chunk_token_budget / chars_per_token)
    n_chunks = (n_tokens // chunk_token_budget) + 1
    overlap_chars = 200
    base_size = (len(content) + (n_chunks - 1) * overlap_chars) // n_chunks
    chunks: list[str] = []
    pos = 0
    while pos < len(content):
        end = min(pos + base_size + overlap_chars, len(content))
        chunks.append(content[pos:end])
        if end >= len(content):
            break
        pos = end - overlap_chars
    return chunks


def _summarize_large_file(params: dict, config: dict) -> str:
    """Multi-agent map-reduce summarization of a potentially large file.

    1. Reads the file (PDF / text / code) — handles big files that would
       overflow Read's normal output.
    2. Estimates token count vs the model's context window.
    3. If it fits → single-shot summary.
       Else → chunks adaptively (number scales with file size), summarizes
       each chunk in parallel via sub-LLM calls (ThreadPoolExecutor with
       up to 8 workers), then a reduce step merges them into one unified
       summary.

    Per-chunk summarization failures are logged inline as
    `[chunk N: error]` markers so one flaky source doesn't sink the
    whole job."""
    from concurrent.futures import ThreadPoolExecutor
    from cheetahclaws.compaction import get_context_limit

    file_path = params.get("file_path", "")
    if not file_path:
        return "Error: missing required parameter 'file_path'"
    focus = params.get("focus", "") or ""

    content = _read_file_for_summary(file_path, config)
    if content.startswith("Error"):
        return content

    model = config.get("model", "")
    model_ctx = get_context_limit(model) or 32768
    chunks = _plan_chunks(content, model_ctx)
    n_chunks = len(chunks)

    p = Path(file_path)
    n_chars = len(content)
    n_tokens_est = _estimate_text_tokens(content)

    if n_chunks == 1:
        summary = _summarize_chunk_via_llm(chunks[0], focus, config, mode="single")
        return (
            f"Summary of `{p.name}` (single-shot, ~{n_tokens_est:,} tokens "
            f"in {n_chars:,} chars; model context {model_ctx:,}):\n\n"
            f"{summary}"
        )

    # Map: parallel summarize chunks
    max_workers = min(n_chunks, 8)
    chunk_summaries: list[str | None] = [None] * n_chunks

    def _do_chunk(i_text):
        i, text = i_text
        return i, _summarize_chunk_via_llm(
            text, focus, config, mode="map",
            chunk_idx=i + 1, total_chunks=n_chunks,
        )

    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        for i, summary_text in ex.map(_do_chunk, enumerate(chunks)):
            chunk_summaries[i] = summary_text

    # Reduce only a bounded aggregate of map outputs. The source and every map
    # response are capped independently, so a pathological document cannot
    # build an unbounded in-memory reduce prompt.
    try:
        reduce_cap = int(config.get(
            "summarize_reduce_max_input_chars", _DEFAULT_SUMMARIZE_REDUCE_INPUT_CHARS,
        ))
    except (TypeError, ValueError):
        reduce_cap = _DEFAULT_SUMMARIZE_REDUCE_INPUT_CHARS
    reduce_cap = max(2_000, reduce_cap)
    merged_parts: list[str] = []
    merged_chars = 0
    for i, summary_text in enumerate(chunk_summaries):
        if summary_text is None or merged_chars >= reduce_cap:
            break
        part = f"=== Chunk {i + 1}/{n_chunks} ===\n{summary_text}\n\n"
        remaining = reduce_cap - merged_chars
        merged_parts.append(part[:remaining])
        merged_chars += min(len(part), remaining)
    merged_input = "".join(merged_parts)
    final = _summarize_chunk_via_llm(merged_input, focus, config, mode="reduce")

    return (
        f"Summary of `{p.name}` (multi-agent map-reduce: {n_chunks} chunks, "
        f"~{n_tokens_est:,} tokens in {n_chars:,} chars; model context "
        f"{model_ctx:,}; {max_workers} parallel workers):\n\n{final}"
    )


register_tool(ToolDef(
    name="SummarizeLargeFile",
    schema={
        "name": "SummarizeLargeFile",
        "description": (
            "Summarize a file that may be too large to fit in your context "
            "window. Reads the file (PDF / txt / md / code), splits it "
            "into N chunks adaptive to file size (1 chunk if it fits, "
            "more for larger files within the configured input cap), summarizes each chunk "
            "in parallel via sub-LLM calls (up to 8 workers), then merges "
            "into one unified summary. Use this for papers, books, long "
            "logs, large code files, or any document where Read would "
            "overflow the context window. Returns a single coherent "
            "summary as the tool result."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Absolute path to the file to summarize.",
                },
                "focus": {
                    "type": "string",
                    "description": (
                        "Optional focus area for the summary "
                        "(e.g. 'methodology and benchmarks', 'security risks')."
                    ),
                },
            },
            "required": ["file_path"],
        },
    },
    func=_summarize_large_file,
    read_only=True,
    concurrent_safe=True,
))
